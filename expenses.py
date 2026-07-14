"""
Expenses tracking module.
Parses credit card Excel exports (Max format) and bank screenshots,
filters to a calendar month, and generates income-vs-expenses reports.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import logging

log = logging.getLogger(__name__)

# ── Bank screenshot expenses ───────────────────────────────────────────────

CC_CHARGE_PATTERNS = ("ממקס", "מקס איט", "כרטיס אשראי", "ישראכרט", "לאומי קארד", "isracard", "max it")

def is_credit_card_charge(desc: str) -> bool:
    """True if this bank debit is a credit card monthly charge (already counted via Excel)."""
    d = (desc or "").replace('"', '').replace("'", "").lower()
    return any(p in d for p in CC_CHARGE_PATTERNS)


def categorize_bank_expense(desc: str) -> str:
    """Simple keyword categorization for bank (non-credit-card) debits."""
    d = desc or ""
    if any(p in d for p in ('ני"ע', "ניירות ערך", "SPY", 'קניית נ')):
        return "השקעות וני\"ע"
    if "חיסכון" in d or "חסכון" in d:
        return "חיסכון"
    if "דמי מנוי" in d or "עמל" in d:
        return "עמלות בנק"
    if "משכנתא" in d:
        return "משכנתא"
    if "הלווא" in d:
        return "הלוואות"
    if "העברה" in d or "ביט" in d:
        return "העברות (בנק)"
    return "בנק - אחר"


# ── Parse credit card Excel ────────────────────────────────────────────────

def parse_credit_card_excel(file_path: str) -> list[dict]:
    """
    Parse a Max/Isracard credit card export xlsx.
    Returns list of transactions with: date, name, category, amount, charge_date, notes, tx_type
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    transactions = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_foreign = "חו\"ל" in sheet_name or "מט\"ח" in sheet_name

        # Find header row (contains "תאריך עסקה")
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and row[0] == "תאריך עסקה":
                header_row = i
                break

        if not header_row:
            continue

        # Parse data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row[0]:  # skip empty rows
                continue

            date_str = str(row[0]).strip()
            name = str(row[1] or "").strip()
            category = str(row[2] or "שונות").strip()
            amount = row[5] or 0  # סכום חיוב
            charge_date_str = str(row[9] or "").strip()
            notes = str(row[10] or "").strip()
            tx_type = str(row[4] or "רגילה").strip()

            # Parse dates
            tx_date = _parse_date(date_str)
            charge_date = _parse_date(charge_date_str)

            if tx_date and amount:
                transactions.append({
                    "date": tx_date,
                    "name": name,
                    "category": category,
                    "amount": float(amount),
                    "charge_date": charge_date,
                    "notes": notes,
                    "tx_type": tx_type,
                    "is_foreign": is_foreign,
                    "source": "credit_card",
                })

    wb.close()
    return transactions


def _parse_date(date_str: str) -> datetime | None:
    """Parse date from various formats."""
    if not date_str or date_str == "None":
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue
    return None


# ── Filter to calendar month ──────────────────────────────────────────────

def filter_to_month(transactions: list[dict], year: int, month: int) -> list[dict]:
    """
    Filter transactions to a specific calendar month.
    All transactions filtered by charge date (when money actually left the account).
    """
    filtered = []
    for t in transactions:
        # Use charge date (when money actually left the account)
        check_date = t.get("charge_date") or t["date"]
        if check_date.year == year and check_date.month == month:
            filtered.append(t)
    return filtered


# ── Categorize and summarize ──────────────────────────────────────────────

def summarize_by_category(transactions: list[dict]) -> dict:
    """Group transactions by category and return totals."""
    categories = {}
    for t in transactions:
        cat = t["category"]
        if cat not in categories:
            categories[cat] = {"total": 0, "count": 0, "items": []}
        categories[cat]["total"] += t["amount"]
        categories[cat]["count"] += 1
        categories[cat]["items"].append(t)
    # Sort by total descending
    return dict(sorted(categories.items(), key=lambda x: -x[1]["total"]))


# ── Generate Excel report ─────────────────────────────────────────────────

def generate_report(
    month_key: str,
    expenses: list[dict],
    all_expenses: list[dict],
    income_total: float,
    income_count: int,
    income_details: list[dict],
    output_path: str
) -> str:
    """
    Generate an xlsx report with:
    - Summary: income vs expenses + category comparison to previous month
    - Expenses by category
    - Detailed expense list
    - Filtered out expenses (other months)
    - Income details
    """
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    cat_font = Font(name="Arial", bold=True, size=11)
    cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    money_fmt = '#,##0.00 ₪'
    pct_fmt = '0.0%'
    normal_font = Font(name="Arial", size=10)
    green_font = Font(name="Arial", size=12, bold=True, color="006600")
    red_font = Font(name="Arial", size=12, bold=True, color="CC0000")
    border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    year, mon = month_key.split("-")
    year, mon = int(year), int(mon)
    title = f"דוח הכנסות והוצאות — {month_key}"

    # Calculate previous month data
    if mon == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, mon - 1
    prev_key = f"{prev_year}-{prev_month:02d}"
    prev_expenses = filter_to_month(all_expenses, prev_year, prev_month)
    prev_categories = summarize_by_category(prev_expenses)
    prev_total = sum(t["amount"] for t in prev_expenses)

    # Current month filtered out expenses
    filtered_out = [t for t in all_expenses if t not in expenses]

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "סיכום"
    ws.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)

    expenses_total = sum(t["amount"] for t in expenses)
    balance = income_total - expenses_total
    categories = summarize_by_category(expenses)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Income vs Expenses summary
    ws["A3"] = "הכנסות"
    ws["B3"] = income_total
    ws["A3"].font = green_font
    ws["B3"].font = green_font
    ws["B3"].number_format = money_fmt

    ws["A4"] = "הוצאות"
    ws["B4"] = expenses_total
    ws["A4"].font = red_font
    ws["B4"].font = red_font
    ws["B4"].number_format = money_fmt

    ws["A5"] = "יתרה"
    ws["B5"] = balance
    ws["A5"].font = Font(name="Arial", bold=True, size=12)
    bal_color = "006600" if balance >= 0 else "CC0000"
    ws["B5"].font = Font(name="Arial", bold=True, size=12, color=bal_color)
    ws["B5"].number_format = money_fmt

    # Category breakdown with previous month comparison
    ws["A7"] = "פירוט הוצאות לפי קטגוריה"
    ws["A7"].font = Font(name="Arial", bold=True, size=13)

    headers = ["קטגוריה", "סכום", "מספר עסקאות", "אחוז מסך ההוצאות", f"סכום {prev_key}", f"שינוי"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 9
    for cat, data in categories.items():
        ws.cell(row=row, column=1, value=cat).font = normal_font
        c = ws.cell(row=row, column=2, value=data["total"])
        c.number_format = money_fmt
        c.font = normal_font
        ws.cell(row=row, column=3, value=data["count"]).font = normal_font
        pct_cell = ws.cell(row=row, column=4, value=data["total"] / expenses_total if expenses_total else 0)
        pct_cell.number_format = pct_fmt
        pct_cell.font = normal_font

        # Previous month comparison
        prev_cat_total = prev_categories.get(cat, {}).get("total", 0)
        c_prev = ws.cell(row=row, column=5, value=prev_cat_total)
        c_prev.number_format = money_fmt
        c_prev.font = normal_font

        diff = data["total"] - prev_cat_total
        c_diff = ws.cell(row=row, column=6, value=diff)
        c_diff.number_format = money_fmt
        c_diff.font = Font(name="Arial", size=10, color="CC0000" if diff > 0 else "006600")

        for j in range(1, 7):
            ws.cell(row=row, column=j).border = border
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="סה\"כ").font = cat_font
    c = ws.cell(row=row, column=2, value=expenses_total)
    c.number_format = money_fmt
    c.font = cat_font
    ws.cell(row=row, column=3, value=len(expenses)).font = cat_font
    c_prev_total = ws.cell(row=row, column=5, value=prev_total)
    c_prev_total.number_format = money_fmt
    c_prev_total.font = cat_font
    diff_total = expenses_total - prev_total
    c_diff_total = ws.cell(row=row, column=6, value=diff_total)
    c_diff_total.number_format = money_fmt
    c_diff_total.font = cat_font

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15

    # ── Sheet 2: Expense details ──
    ws2 = wb.create_sheet("פירוט הוצאות")
    ws2.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)

    detail_headers = ["תאריך", "שם בית עסק", "קטגוריה", "סכום", "מקור", "סוג עסקה", "הערות"]
    for j, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, t in enumerate(sorted(expenses, key=lambda x: x.get("charge_date") or x["date"]), 2):
        ws2.cell(row=i, column=1, value=t["date"].strftime("%d/%m/%Y")).font = normal_font
        ws2.cell(row=i, column=2, value=t["name"]).font = normal_font
        ws2.cell(row=i, column=3, value=t["category"]).font = normal_font
        c = ws2.cell(row=i, column=4, value=t["amount"])
        c.number_format = money_fmt
        c.font = normal_font
        source_label = "חשבון בנק" if t.get("source") == "bank" else "אשראי"
        ws2.cell(row=i, column=5, value=source_label).font = normal_font
        ws2.cell(row=i, column=6, value=t["tx_type"]).font = normal_font
        ws2.cell(row=i, column=7, value=t["notes"]).font = normal_font
        for j in range(1, 8):
            ws2.cell(row=i, column=j).border = border

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 15
    ws2.column_dimensions["E"].width = 13
    ws2.column_dimensions["F"].width = 15
    ws2.column_dimensions["G"].width = 30

    # ── Sheet 3: Filtered out expenses ──
    ws_filtered = wb.create_sheet("הוצאות שסוננו")
    ws_filtered.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)

    filter_headers = ["תאריך עסקה", "תאריך חיוב", "שם בית עסק", "קטגוריה", "סכום", "מקור", "סיבת סינון"]
    for j, h in enumerate(filter_headers, 1):
        cell = ws_filtered.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, t in enumerate(sorted(filtered_out, key=lambda x: x.get("charge_date") or x["date"]), 2):
        ws_filtered.cell(row=i, column=1, value=t["date"].strftime("%d/%m/%Y")).font = normal_font
        charge = t.get("charge_date")
        ws_filtered.cell(row=i, column=2, value=charge.strftime("%d/%m/%Y") if charge else "").font = normal_font
        ws_filtered.cell(row=i, column=3, value=t["name"]).font = normal_font
        ws_filtered.cell(row=i, column=4, value=t["category"]).font = normal_font
        c = ws_filtered.cell(row=i, column=5, value=t["amount"])
        c.number_format = money_fmt
        c.font = normal_font
        source_label = "חשבון בנק" if t.get("source") == "bank" else "אשראי"
        ws_filtered.cell(row=i, column=6, value=source_label).font = normal_font
        charge_month = charge.strftime("%Y-%m") if charge else "?"
        ws_filtered.cell(row=i, column=7, value=f"חודש {charge_month}").font = normal_font
        for j in range(1, 8):
            ws_filtered.cell(row=i, column=j).border = border

    ws_filtered.column_dimensions["A"].width = 14
    ws_filtered.column_dimensions["B"].width = 14
    ws_filtered.column_dimensions["C"].width = 30
    ws_filtered.column_dimensions["D"].width = 20
    ws_filtered.column_dimensions["E"].width = 15
    ws_filtered.column_dimensions["F"].width = 13
    ws_filtered.column_dimensions["G"].width = 15

    # ── Sheet 4: Income details ──
    ws3 = wb.create_sheet("פירוט הכנסות")
    ws3.sheet_properties.sheetView = openpyxl.worksheet.views.SheetView(rightToLeft=True)

    inc_headers = ["תאריך", "לקוח", "סכום"]
    for j, h in enumerate(inc_headers, 1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, inc in enumerate(income_details, 2):
        ws3.cell(row=i, column=1, value=inc.get("bank_date", "")).font = normal_font
        ws3.cell(row=i, column=2, value=inc.get("customer_name", "")).font = normal_font
        c = ws3.cell(row=i, column=3, value=inc.get("amount", 0))
        c.number_format = money_fmt
        c.font = normal_font
        for j in range(1, 4):
            ws3.cell(row=i, column=j).border = border

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 15

    # Save
    wb.save(output_path)
    log.info(f"Report saved to {output_path}")
    return output_path
